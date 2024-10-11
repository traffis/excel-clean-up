import openpyxl

# open billing workbook and clean up sheet
workbook = openpyxl.load_workbook("Billing.xlsx") # needs this file to perform operations on
clean_up_sheet = workbook["Clean Up"]

def update_tickets(mysoft_tickets, mysoft_work_orders):
    """Removes tickets from Clean Up sheet and adds tickets from MySoft"""
    current_tickets = []

    # remove tickets from sheet that aren't in mysoft
    i = 2
    while i <= clean_up_sheet.max_row:
        if str(clean_up_sheet.cell(row=i, column=2).value) not in mysoft_work_orders:
            print("Removed WO: " + str(clean_up_sheet.cell(row=i, column=2).value))
            clean_up_sheet.delete_rows(i, 1)
        else:
            current_tickets.append(str(clean_up_sheet.cell(row=i, column=2).value))
            i += 1

    # add tickets from mysoft to sheet
    for mysoft_work_order in mysoft_work_orders:
        if mysoft_work_order not in current_tickets and mysoft_tickets[mysoft_work_order].isdigit():
            clean_up_sheet.append([int(mysoft_tickets[mysoft_work_order]), int(mysoft_work_order)])
            print("Added WO: " + str(mysoft_work_order))

        elif mysoft_work_order not in current_tickets and not mysoft_tickets[mysoft_work_order].isdigit():
            clean_up_sheet.append([mysoft_tickets[mysoft_work_order], int(mysoft_work_order)])
            print("Added WO: " + str(mysoft_work_order))

    workbook.save("Billing.xlsx")
